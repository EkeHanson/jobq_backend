from rest_framework import viewsets, generics, permissions, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.views import APIView
from django.db.models import Count
from django.utils import timezone
from django.db.models import Q
from django.conf import settings
import openpyxl
from datetime import datetime

from .models import Application, StatusHistory, Interview, BulkImportTask
from .serializers import ApplicationSerializer, StatusHistorySerializer, InterviewSerializer, BulkImportTaskSerializer


class ApplicationViewSet(viewsets.ModelViewSet):
    queryset = Application.objects.select_related('user').prefetch_related('history').all().order_by('-created_at')
    serializer_class = ApplicationSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        qs = super().get_queryset()
        user = self.request.user
        if not self.request.user.is_staff:
            qs = qs.filter(user=user)
        
        # Filter by archived status - show all by default
        archived = self.request.query_params.get('archived')
        if archived is not None:
            qs = qs.filter(archived=archived.lower() == 'true')
        # No filter param = show all (both active and archived)
        
        # Filter out soft-deleted applications by default
        # Include deleted if explicitly requested
        include_deleted = self.request.query_params.get('include_deleted', 'false').lower() == 'true'
        if not include_deleted:
            qs = qs.filter(deleted_at__isnull=True)
        
        return qs
    
    @action(detail=False, methods=['get'])
    def stats(self, request):
        """Get application statistics for current user"""
        qs = Application.objects.filter(user=request.user)
        total = qs.count()
        
        # Use Django aggregation for efficient status counting - O(1) instead of O(n)
        status_counts = qs.values('status').annotate(count=Count('id'))
        by_status = {item['status'] or 'unknown': item['count'] for item in status_counts}
        
        total_companies = qs.values('company_name').distinct().count()
        
        interview_rate = (by_status.get('interview', 0) / total * 100) if total else 0
        offer_rate = (by_status.get('offer', 0) / total * 100) if total else 0
        response_rate = (
            (by_status.get('interview', 0) + by_status.get('offer', 0) + by_status.get('rejected', 0))
            / total * 100
        ) if total else 0
        
        stats = {
            'total': total,
            'total_companies': total_companies,
            'by_status': by_status,
            'response_rate': response_rate,
            'interview_rate': interview_rate,
            'offer_rate': offer_rate,
            'archived_count': qs.filter(archived=True).count(),
        }
        return Response(stats)

    @action(detail=True, methods=['post'])
    def archive(self, request, pk=None):
        """Archive an application"""
        application = self.get_object()
        application.archived = True
        application.archived_at = timezone.now()
        application.save()
        return Response({'status': 'archived', 'archived_at': application.archived_at})

    @action(detail=True, methods=['post'])
    def unarchive(self, request, pk=None):
        """Unarchive an application"""
        application = self.get_object()
        application.archived = False
        application.archived_at = None
        application.save()
        return Response({'status': 'unarchived'})

    @action(detail=True, methods=['post'])
    def soft_delete(self, request, pk=None):
        """Soft delete an application"""
        application = self.get_object()
        application.deleted_at = timezone.now()
        application.save()
        return Response({'status': 'deleted', 'deleted_at': application.deleted_at})

    @action(detail=True, methods=['post'])
    def restore(self, request, pk=None):
        """Restore a soft-deleted application"""
        application = self.get_object()
        application.deleted_at = None
        application.save()
        return Response({'status': 'restored'})


class StatusHistoryView(generics.ListAPIView):
    serializer_class = StatusHistorySerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        application_id = self.kwargs['pk']
        return StatusHistory.objects.filter(application_id=application_id)


class FollowUpsView(generics.ListAPIView):
    """View for getting follow-up reminders"""
    serializer_class = ApplicationSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        today = timezone.now().date()
        # Get applications with follow-up dates in the future or today
        return Application.objects.filter(
            user=self.request.user,
            follow_up_date__isnull=False,
            follow_up_date__lte=today + timezone.timedelta(days=7),  # Up to 7 days ahead
            archived=False,
            deleted_at__isnull=True
        ).order_by('follow_up_date')

    @action(detail=False, methods=['post'])
    def mark_sent(self, request):
        """Mark a follow-up as sent"""
        application_id = request.data.get('application_id')
        try:
            application = Application.objects.get(id=application_id, user=request.user)
            application.follow_up_sent = True
            application.save()
            return Response({'status': 'marked_as_sent'})
        except Application.DoesNotExist:
            return Response({'error': 'Application not found'}, status=status.HTTP_404_NOT_FOUND)


class InterviewViewSet(viewsets.ModelViewSet):
    """ViewSet for managing interviews"""
    serializer_class = InterviewSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        return Interview.objects.filter(
            application__user=self.request.user
        ).select_related('application').order_by('-interview_date', '-created_at')
    
    def perform_create(self, serializer):
        application_id = self.request.data.get('application')
        application = Application.objects.get(id=application_id, user=self.request.user)
        serializer.save(application=application)
    
    @action(detail=True, methods=['post'])
    def update_outcome(self, request, pk=None):
        """Update interview outcome"""
        interview = self.get_object()
        outcome = request.data.get('outcome')
        if outcome in ['pending', 'passed', 'failed', 'cancelled']:
            interview.outcome = outcome
            interview.save()
            return Response({'status': 'outcome_updated', 'outcome': outcome})
        return Response({'error': 'Invalid outcome'}, status=status.HTTP_400_BAD_REQUEST)


class ApplicationByStatusView(generics.ListAPIView):
    """View for getting applications grouped by status (for Kanban)"""
    serializer_class = ApplicationSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        return Application.objects.filter(
            user=self.request.user,
            archived=False,
            deleted_at__isnull=True
        ).prefetch_related('history', 'interviews').order_by('-created_at')


class BulkImportView(generics.CreateAPIView):
    """View for initiating bulk import of applications from Excel"""
    queryset = BulkImportTask.objects.all()
    serializer_class = BulkImportTaskSerializer
    permission_classes = [permissions.IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]
    
    def create(self, request, *args, **kwargs):
        file_obj = request.FILES.get('file')
        
        if not file_obj:
            return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Validate file type
        allowed_extensions = ['.xlsx', '.xls']
        file_name = file_obj.name.lower()
        if not any(file_name.endswith(ext) for ext in allowed_extensions):
            return Response({'error': 'Invalid file type. Please upload an Excel file (.xlsx or .xls)'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        # Create the bulk import task
        task = BulkImportTask.objects.create(
            user=request.user,
            file_name=file_obj.name,
            status='pending'
        )
        
        # Start processing in background (synchronously for now, but could be async)
        try:
            task.status = 'processing'
            task.save()
            
            # Process the file
            result = process_bulk_import(task, file_obj)
            
            return Response({
                'task_id': task.task_id,
                'status': task.status,
                'total_rows': task.total_rows,
                'successful_rows': task.successful_rows,
                'failed_rows': task.failed_rows,
                'message': f'Successfully imported {task.successful_rows} applications'
            }, status=status.HTTP_201_CREATED)
            
        except Exception as e:
            task.status = 'failed'
            task.errors = [{'row': 0, 'error': str(e)}]
            task.save()
            return Response({
                'task_id': task.task_id,
                'status': 'failed',
                'error': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class BulkImportStatusView(generics.RetrieveAPIView):
    """View for checking bulk import task status"""
    queryset = BulkImportTask.objects.all()
    serializer_class = BulkImportTaskSerializer
    permission_classes = [permissions.IsAuthenticated]
    lookup_field = 'task_id'
    
    def get_queryset(self):
        return BulkImportTask.objects.filter(user=self.request.user)


def process_bulk_import(task, file_obj):
    """Process the Excel file and create applications"""
    wb = openpyxl.load_workbook(file_obj)
    ws = wb.active
    
    # Get headers from first row
    headers = []
    for cell in ws[1]:
        headers.append(cell.value)
    
    # Map headers to model fields
    field_mapping = {
        'job_title': 'job_title',
        'company_name': 'company_name',
        'company': 'company_name',
        'status': 'status',
        'applied_date': 'applied_date',
        'applied': 'applied_date',
        'date_applied': 'applied_date',
        'deadline': 'deadline',
        'notes': 'notes',
        'source': 'source',
        'follow_up_date': 'follow_up_date',
        'follow_up': 'follow_up_date',
        'description': 'description',
        'requirements': 'requirements',
    }
    
    # Valid status values
    valid_statuses = ['saved', 'applied', 'assessment', 'interview', 'offer', 'rejected', 'accepted', 'withdrawn']
    valid_sources = ['linkedin', 'indeed', 'jobberman', 'glassdoor', 'company_website', 'referral', 'recruiter', 'other']
    
    total_rows = ws.max_row - 1  # Exclude header row
    task.total_rows = total_rows
    task.save()
    
    errors = []
    successful = 0
    
    # Process each row
    for row_num in range(2, ws.max_row + 1):
        row_data = {}
        row_errors = []
        
        for col_idx, header in enumerate(headers, start=1):
            if header:
                header_lower = str(header).strip().lower()
                field_name = field_mapping.get(header_lower, header_lower)
                cell_value = ws.cell(row=row_num, column=col_idx).value
                
                # Parse dates
                if field_name in ['applied_date', 'deadline', 'follow_up_date'] and cell_value:
                    if isinstance(cell_value, datetime):
                        row_data[field_name] = cell_value.date()
                    elif isinstance(cell_value, str):
                        try:
                            row_data[field_name] = datetime.strptime(cell_value, '%Y-%m-%d').date()
                        except ValueError:
                            try:
                                row_data[field_name] = datetime.strptime(cell_value, '%m/%d/%Y').date()
                            except ValueError:
                                row_errors.append(f'Invalid date format for {header}: {cell_value}')
                    continue
                
                # Validate status
                if field_name == 'status' and cell_value:
                    if cell_value.lower() not in valid_statuses:
                        row_errors.append(f'Invalid status: {cell_value}. Must be one of: {valid_statuses}')
                    else:
                        row_data[field_name] = cell_value.lower()
                    continue
                
                # Validate source
                if field_name == 'source' and cell_value:
                    if cell_value.lower() not in valid_sources:
                        row_errors.append(f'Invalid source: {cell_value}. Must be one of: {valid_sources}')
                    else:
                        row_data[field_name] = cell_value.lower()
                    continue
                
                if cell_value is not None:
                    row_data[field_name] = cell_value
        
        # Validate required fields
        if not row_data.get('job_title'):
            row_errors.append('Job title is required')
        if not row_data.get('company_name'):
            row_errors.append('Company name is required')
        
        # Set default status if not provided
        if 'status' not in row_data:
            row_data['status'] = 'saved'
        
        if row_errors:
            errors.append({'row': row_num, 'errors': row_errors})
            task.failed_rows += 1
        else:
            try:
                Application.objects.create(
                    user=task.user,
                    **row_data
                )
                successful += 1
                task.successful_rows += 1
            except Exception as e:
                errors.append({'row': row_num, 'errors': [str(e)]})
                task.failed_rows += 1
        
        task.processed_rows += 1
        task.save()
    
    # Mark task as completed
    task.status = 'completed'
    task.completed_at = timezone.now()
    task.errors = errors
    task.save()
    
    return {
        'total': total_rows,
        'successful': successful,
        'failed': task.failed_rows
    }
